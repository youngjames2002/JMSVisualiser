import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO, StringIO
import msal
import requests
import json
import re

@st.cache_data(show_spinner=True)
def get_machine_schedule_labels() -> pd.DataFrame:
    TENANT_ID = st.secrets["sharepoint"]["TENANT_ID"]
    CLIENT_ID = st.secrets["sharepoint"]["CLIENT_ID"]
    CLIENT_SECRET = st.secrets["sharepoint"]["CLIENT_SECRET"]
    PLAN_ID = "OMiGy-Z9OE2SZd14RwvvfJYAAeYC"

    OPERATION_CATEGORIES = {
        "category1": "CNC Milling",
        "category2": "Csking/Drilling",
        "category3": "CNC Turning",
        "category4": "Manual Turning",
        "category5": "After Weld Machining",
        "category16": "Flanges",
    }
    SITE_CATEGORIES = {
        "category19": "Kilrea",
        "category23": "Ballymena",
    }

    app = msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET
    )
    token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

    if "access_token" not in token:
        st.error(f"Planner auth failed: {token.get('error_description')}")
        return pd.DataFrame()

    headers = {"Authorization": f"Bearer {token['access_token']}"}

    tasks_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/planner/plans/{PLAN_ID}/tasks",
        headers=headers
    )
    tasks_resp.raise_for_status()
    tasks = tasks_resp.json()["value"]

    rows = []
    for task in tasks:
        match = re.match(r"(SO-\d+)", task["title"])
        if not match:
            continue

        so = match.group(1)
        applied = task.get("appliedCategories", {})

        site = "No Site Assigned"
        for k, v in SITE_CATEGORIES.items():
            if k in applied:
                site = v
                break

        operations = [OPERATION_CATEGORIES[k] for k in applied if k in OPERATION_CATEGORIES]
        for operation in operations:
            rows.append({"S.O. No.": so, "Operation": operation, "Site": site})

    return pd.DataFrame(rows)

@st.cache_data(ttl=1500)
def get_statii_session_token() -> str:
    BASE_URL     = st.secrets["statii"]["BASE_URL"]
    CLIENT_SECRET = st.secrets["statii"]["CLIENT_SECRET"]
    CLIENT_ID     = st.secrets["statii"]["CLIENT_ID"]
    # Docs: username=Client Secret, password=Client Id — try swapping if 401 persists
    response = requests.get(f"{BASE_URL}/auth", auth=(CLIENT_ID, CLIENT_SECRET))
    if not response.ok:
        print("Auth failed:", response.status_code, response.text)
        response.raise_for_status()
    return response.json()["ResponseBody"]["data"]["session"]

@st.cache_data(show_spinner=True)
def statii_paint_data():
    BASE_URL = st.secrets["statii"]["BASE_URL"]
    token = get_statii_session_token()
    response = requests.get(
        f"{BASE_URL}/report/sales_order_lines",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
        params={"filters": json.dumps({
            "live": True,
            "specification": [
                {"pattern": "RAL"},
                {"pattern": "prime"},
                {"pattern": "paint"},
                {"pattern": "rl"},
                {"pattern": "bs"},
            ],
        })},
    )
    response.raise_for_status()
    data = response.json()["ResponseBody"]["data"]
    return data

@st.cache_data(show_spinner=True)
def statii_galv_data():
    BASE_URL = st.secrets["statii"]["BASE_URL"]
    token = get_statii_session_token()
    response = requests.get(
        f"{BASE_URL}/report/sales_order_lines",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
        params={"filters": json.dumps({
            "live": True,
            "specification": {"pattern": "Galv"},
        })},
    )
    response.raise_for_status()
    data = response.json()["ResponseBody"]["data"]
    return data
    
@st.cache_data(show_spinner=True)
def statii_bundle_jobs(operation):
    BASE_URL     = st.secrets["statii"]["BASE_URL"]
    token = get_statii_session_token()
    response = requests.get(
        f"{BASE_URL}/report/scheduling",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
        params={"filters": json.dumps({"live": True})},
    )
    data = response.json()["ResponseBody"]["data"]
    df = pd.DataFrame(data["rows"], columns=data["columns"])
    return df[df["operation"] == operation]

def clean_statii_bundle_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # map column names — confirmed against /report/scheduling API response
    df = df.rename(columns={
        "number":         "Number",
        "customer":       "Customer",
        "so_no":          "S.O. No.",
        "hours_plan":     "Hours Plan",
        "date_requested": "Date Requested",
        "operation":      "Operation",
    })

    df["Hours Plan"] = pd.to_numeric(df["Hours Plan"], errors="coerce").fillna(0)
    df["Time Planned"] = df["Hours Plan"].apply(format_hours)
    df["Date Requested"] = pd.to_datetime(df["Date Requested"], errors="coerce")

    # apply company grouping (adds Customer Grouped)
    df = apply_company_grouping(df)

    # site logic: Bamford → Ballymena, everyone else → Kilrea
    df["Site"] = (
        df["Customer Grouped"]
        .str.contains("BAMFORD", case=False, na=False)
        .map({True: "Ballymena", False: "Kilrea"})
    )

    # week ending: next Friday after Date Requested, formatted to match existing schedule tools
    df["Week Ending"] = (
        df["Date Requested"] + pd.offsets.Week(weekday=4)
    ).dt.strftime("%d/%m/%Y")

    return df

@st.cache_data(show_spinner=True)
def statii_completed_jobs():
    BASE_URL     = st.secrets["statii"]["BASE_URL"]
    token = get_statii_session_token()
    response = requests.get(
        f"{BASE_URL}/report/scheduling",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
        params={"filters": json.dumps({"status": "Complete"})},
    )
    response.raise_for_status()
    data = response.json()["ResponseBody"]["data"]
    return data

def clean_paint_data_from_api(api_response: dict) -> pd.DataFrame:
    df = pd.DataFrame(api_response["rows"], columns=api_response["columns"])

    # Filter out non-painted customers
    # this customer filter can defo be done on the API call to reduce the amount of records pulled
    df = df[~df["customer"].str.contains("Bamford|Wright|Cunningham", case=False, na=False)] 
    # Filter for paint-related specifications
    df = df[df["specification"].str.contains(r"\bRAL\b|\bprime\b|\bpaint\b|\bpainted\b|\bprimed\b|\brl\b|\bbs\b", case=False, na=False)]

    # Date handling — API returns ISO 8601 so no dayfirst needed
    df["date_promised"] = pd.to_datetime(df["date_promised"], errors="coerce")
    df = df.dropna(subset=["date_promised"])
    df["date_promised"] = df["date_promised"] - pd.Timedelta(days=2)
    df["Week Due"] = df["date_promised"].dt.to_period("W-FRI").apply(lambda r: r.end_time)
    current_week = pd.Timestamp.today().to_period("W-FRI").end_time
    df = df[df["Week Due"] >= current_week]
    df["Week Label"] = df["Week Due"].dt.strftime("%d %b")
    df = df.sort_values("Week Due", ascending=True)

    df = df.rename(columns={
        "number": "Line No",
        "customer": "Customer",
        "specification": "Specification",
        "price": "Price",
        "date_promised": "Date Promised",
        "description" : "Description"
    })

    df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0)

    return df[["Line No", "Customer", "Specification", "Description", "Price", "Date Promised", "Week Due", "Week Label"]]

def clean_galv_data(api_response: dict) -> pd.DataFrame:
    df = pd.DataFrame(api_response["rows"], columns=api_response["columns"])

    # Filter out non-painted customers
    # this customer filter can defo be done on the API call to reduce the amount of records pulled
    df = df[~df["customer"].str.contains("Bamford|Wright|Cunningham", case=False, na=False)] 
    # Filter for paint-related specifications
    df = df[df["specification"].str.contains("Galv", case=False, na=False)]

    # Date handling — API returns ISO 8601 so no dayfirst needed
    df["date_promised"] = pd.to_datetime(df["date_promised"], errors="coerce")
    df = df.dropna(subset=["date_promised"])
    df["date_promised"] = df["date_promised"] - pd.Timedelta(days=2)
    df["Week Due"] = df["date_promised"].dt.to_period("W-FRI").apply(lambda r: r.end_time)
    current_week = pd.Timestamp.today().to_period("W-FRI").end_time
    df = df[df["Week Due"] >= current_week]
    df["Week Label"] = df["Week Due"].dt.strftime("%d %b")
    df = df.sort_values("Week Due", ascending=True)

    df = df.rename(columns={
        "number": "Line No",
        "customer": "Customer",
        "specification": "Specification",
        "price": "Price",
        "date_promised": "Date Promised",
        "description" : "Description"
    })

    df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0)

    return df[["Line No", "Customer", "Specification", "Description", "Price", "Date Promised", "Week Due", "Week Label"]]

@st.cache_data(show_spinner=True)
def download_excel_from_sharepoint(site_name: str, file_path:str) -> BytesIO:
    # download from sharepoint and return bytesIO object

    TENANT_ID = st.secrets["sharepoint"]["TENANT_ID"]
    CLIENT_ID = st.secrets["sharepoint"]["CLIENT_ID"]
    CLIENT_SECRET = st.secrets["sharepoint"]["CLIENT_SECRET"]
    SHAREPOINT_SITE = st.secrets["sharepoint"]["SHAREPOINT_SITE"]

    AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
    SCOPE = ["https://graph.microsoft.com/.default"]

    app = msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET
    )  
    # debug st.json(token)  # should contain "access_token" if successful

    token = app.acquire_token_for_client(scopes=SCOPE)
    if "access_token" not in token:
        st.error("Authentication failed")
        return None

    headers = {"Authorization": f"Bearer {token['access_token']}"} 

    # Get SharePoint site ID
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE}:/sites/{site_name}:/"
    site_response = requests.get(site_url, headers=headers)
    # debug
    # st.write("Site lookup status:", site_response.status_code)
    # st.write(site_response.json())

    if site_response.status_code != 200:
        st.error("Site lookup failed")
        return None

    site_id = site_response.json()["id"]

    # # debug
    # drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    # drives_response = requests.get(drives_url, headers=headers)
    # st.write(drives_response.json())

    # Download the file
    file_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path}:/content"
    file_response = requests.get(file_url, headers=headers)
    file_response.raise_for_status()

    return BytesIO(file_response.content)

@st.cache_data(show_spinner=True)
def load_data_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/JMS Master Schedule/testAutomation/bundleStagingSheet.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    df = pd.read_excel(bytes_io)

    df["Earliest Process Date"] = pd.to_datetime(
        df["Earliest Process Date"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )
    df = apply_company_grouping(df)
    return df

@st.cache_data(show_spinner=True)
def load_data_completed_jobs(resource):
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Admin/completed_jobs_weld_saw_machining.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()

    sheet_map = {
        "weld": "weldTable",
        "saw": "sawTable",
        "machine": "machineTable"
    }

    sheet_name = sheet_map.get(resource)
    if sheet_name is None:
        st.error(f"Unknown resource: '{resource}'. Expected one of: {list(sheet_map.keys())}")
        return pd.DataFrame()

    df = pd.read_excel(bytes_io, sheet_name=sheet_name)

    return df
    
@st.cache_data(show_spinner=True)
def load_data_Bmena_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Paint Schedule/Bmena Finishing Schedule.xlsm"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    wb = load_workbook(filename=bytes_io, data_only=True)
    sheet = wb["Schedule"]
    lookup_table = sheet.tables["Table1"]
    data = sheet[lookup_table.ref]

    # Convert table to DataFrame
    rows_list = [[cell.value for cell in row] for row in data]
    df = pd.DataFrame(rows_list[1:], columns=rows_list[0])
    return df

@st.cache_data(show_spinner=True)
def load_data_ncr_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/NCR Log/NCR Log.xlsm"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed
    
    wb = load_workbook(filename=bytes_io, data_only=True)
    sheet = wb["1 - Non-Conformance Log"]
    lookup_table = sheet.tables["Table1"]
    data = sheet[lookup_table.ref]
    df = table_to_df(data)

    # fix dates
    df["Date"] = pd.to_datetime(
        df["Date"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )

    return df

@st.cache_data(show_spinner=True)
def load_data_weld_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Admin/Welding Schedule Teams Tool.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    df = pd.read_excel(bytes_io)

    df["Date Requested"] = pd.to_datetime(
        df["Date Requested"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )
    df.columns = df.columns.str.strip()
    df = apply_company_grouping(df)
    return df

@st.cache_data(show_spinner=True)
def load_data_machine_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Admin/Machining Schedule Teams Tool.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    df = pd.read_excel(bytes_io)

    df["Date Requested"] = pd.to_datetime(
        df["Date Requested"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )
    df.columns = df.columns.str.strip()
    df = apply_company_grouping(df)
    return df

@st.cache_data(show_spinner=True)
def load_data_saw_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Admin/Saw Schedule Teams Tool.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    df = pd.read_excel(bytes_io)

    df["Date Requested"] = pd.to_datetime(
        df["Date Requested"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )
    df.columns = df.columns.str.strip()
    df = apply_company_grouping(df)
    return df

@st.cache_data(show_spinner=True)
def load_data_rubber_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/Admin/Rubber Lining Teams Tool.xlsx"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed

    df = pd.read_excel(bytes_io)

    df["Date Requested"] = pd.to_datetime(
        df["Date Requested"],
        format="mixed",
        dayfirst=True,
        errors="coerce"
    )
    df.columns = df.columns.str.strip()
    df = apply_company_grouping(df)
    return df

def table_to_df(data):
    rows_list=[]

    for row in data:
        cols=[]
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)

    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
    return df

@st.cache_data(show_spinner=True)
def load_so_sp():
    bytes_io = download_excel_from_sharepoint(
        site_name="JMSEngineeringTeam",
        file_path="JMS Engineering Team SharePoint/NCR Log/ALL SALES ORDERS.csv"
    )
    if bytes_io is None:
        return pd.DataFrame()  # return empty DataFrame if download failed
    
    df = pd.read_csv(BytesIO(bytes_io.getvalue()))
    return df

@st.cache_data(show_spinner=True)
def load_so_statii():
    BASE_URL     = st.secrets["statii"]["BASE_URL"]
    token = get_statii_session_token()
    response = requests.get(
        f"{BASE_URL}/report/sales_orders",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        },
    )
    response.raise_for_status()
    data = response.json()["ResponseBody"]["data"]
    return pd.DataFrame(data["rows"], columns=data["columns"])


def apply_company_grouping(df):
    df=df.copy()
    df["Customer Grouped"] = df["Customer"].str.upper().str.strip()

    COMPANY_KEYWORDS = [
        "BAMFORD",
        "CDE",
        "TOBERMORE",
        "FARLOW",
        "SANDVIK",
        "CROSSLAND",
        "WRIGHTBUS"
    ]

    for keyword in COMPANY_KEYWORDS:
        df.loc[
            df["Customer Grouped"].str.contains(keyword, na=False),
            "Customer Grouped"
        ] = keyword

    return df

def load_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


def apply_filters(df, late_select, incomplete_only, selected_customers, selected_machines, bundle_search, folding_required):
    filtered_df = df.copy()

    # Late Filter -- Calendar Week
    today = pd.Timestamp.today().normalize()
    current_week = today.to_period("W")

    due_dates = filtered_df["Earliest Process Date"]
    due_weeks = due_dates.dt.to_period("W")

    # Status masks (calendar-based)
    late_mask = due_dates < today
    week_mask = (due_weeks == current_week) & (due_dates >= today)
    future_mask = due_weeks > current_week

    status_mask = False

    if "Late" in late_select:
        status_mask |= late_mask

    if "Due This Week" in late_select:
        status_mask |= week_mask

    if "Due in Future" in late_select:
        status_mask |= future_mask

    if late_select:
        filtered_df = filtered_df[status_mask]    

    # Customer filter
    filtered_df = filtered_df[
        filtered_df["Customer Grouped"].isin(selected_customers)
    ]

    # Machine filter
    filtered_df = filtered_df[
        filtered_df["Machine"].isin(selected_machines)
    ]

    # Incomplete filter
    if incomplete_only:
        filtered_df = filtered_df[
            filtered_df["Completed?"] == "No"
        ]

    # bundle search
    if bundle_search:
        filtered_df = filtered_df[
            filtered_df["Bundle/Job"]
            .astype(str)
            .str.contains(bundle_search, case=False, na=False)
    ]
        
    # folding toggle
    if folding_required:
        filtered_df = filtered_df[filtered_df["Folding Required?"] == "Yes"]

    return filtered_df

def bmena_finishing_filters(df):
    # date filter
    # get today
    today = pd.Timestamp.today().normalize()

    # Ensure datetime
    df["Finish Required Week Ending"] = pd.to_datetime(
        df["Finish Required Week Ending"], errors="coerce"
    )

    # Find the nearest date AFTER today
    future_dates = df.loc[
        df["Finish Required Week Ending"] > today,
        "Finish Required Week Ending"
    ]
    if not future_dates.empty:
        cutoff_date = future_dates.min()
    else:
        # If no future dates exist, use max date in column
        cutoff_date = df["Finish Required Week Ending"].max()

    # Filter rows with date <= cutoff
    filtered_df = df[
        df["Finish Required Week Ending"] <= cutoff_date
    ]  

    # blank filters
    # only records with date delivered AND supplier AND comments blanked
    filtered_df = filtered_df[filtered_df["Date Delivered"].isna()]
    filtered_df = filtered_df[filtered_df["Supplier"].isna()]
    filtered_df = filtered_df[filtered_df["Comments"].isna()]

    return filtered_df
    
def parse_paint_data(raw_data):
    try: 
        df = pd.read_csv(
            StringIO(raw_data), sep="\t", engine="python", quotechar='"', skip_blank_lines=True
        )
    except Exception as e:
        st.error(f"Cannot read data - ensure data is input from statii correctly - {e}")
        st.stop()

    return df

def clean_paint_data(df):
    # filter to remove customers who dont get painted
    df = df[~df["Customer"].str.contains("Bamford|Wright|Cunningham", case=False, na=False)]
    # filter for specifications that are paint
    df = df[df["Specification"].str.contains(r"\bRAL\b|\bprime\b|\bpaint\b", case=False, na=False)]

    # add week column and sort by that
    df["Date Promised"] = pd.to_datetime(
        df["Date Promised"], dayfirst=True, errors="coerce"
    )
    df = df.dropna(subset=["Date Promised"])
    df["Date Promised"] = df["Date Promised"] - pd.Timedelta(days=2)   # paint date is 2 days before so date
    df["Week Due"] = df["Date Promised"].dt.to_period("W-FRI").apply(lambda r: r.end_time)
    current_week = pd.Timestamp.today().to_period("W-FRI").end_time
    df = df[df["Week Due"] >= current_week]
    df["Week Label"] = df["Week Due"].dt.strftime("%d %b")
    df = df.sort_values("Week Due", ascending=True)

    return df
    
def clean_weld_saw_machine_data(df):
    clean_df = df.copy()
    #strip columns we dont use
    clean_df.drop(['PlannerDueDate', 'Task Description', 'PlannerTaskID', 'PlannerCreated', 'CreatedOn'], axis=1, inplace=True)
    #add site logic - bamford = bmena, other = kilrea
    clean_df["Site"] = clean_df["Customer Grouped"].str.contains("BAMFORD", case=False, na=False).map({True: "Ballymena", False: "Kilrea"})
    #add week ending logic
    clean_df["Week Ending"] = (
        pd.to_datetime(clean_df["Date Requested"]) + pd.offsets.Week(weekday=4)
    ).dt.strftime("%d/%m/%Y")
    return clean_df

def clean_flat_data(df):
    clean_df = df.copy()
    # strip unnescessary columns
    clean_df.drop(['bundle-ID', 'Assign to:', 'Date Added'], axis=1, inplace=True)
    # only non completed jobs
    clean_df=clean_df[clean_df['Completed?'] != "Yes"]
    # add week ending logic 
    clean_df["Week Ending"] = (
        pd.to_datetime(clean_df["Earliest Process Date"]) + pd.offsets.Week(weekday=4)
    ).dt.strftime("%d/%m/%Y")
    # add site logic
    clean_df["Site"] = clean_df["Machine"].str.contains("REGIUS", case=False, na=False).map({True: "Ballymena", False: "Kilrea"})
    # clean up hours for visuals
    clean_df["Hours"] = clean_df["Estimated Bundle Time (Hours)"].apply(format_hours)
    return clean_df

def clean_fold_data(df):
    clean_df = df.copy()
    # strip unnescessary columns
    clean_df.drop(['bundle-ID', 'Assign to:', 'Date Added'], axis=1, inplace=True)
    # only non completed jobs
    clean_df=clean_df[clean_df['Fold Completed?'] != "Yes"]
    # add week ending logic 
    clean_df["Week Ending"] = (
        pd.to_datetime(clean_df["Earliest Fold Date"]) + pd.offsets.Week(weekday=4)
    ).dt.strftime("%d/%m/%Y")
    # rename fold site to Site
    clean_df = clean_df.rename(columns={"Fold Site": "Site"})
    # clean up hours for visuals
    clean_df["Hours"] = clean_df["Estimated Fold Time (Hours)"].apply(format_hours)
    return clean_df

def remove_completed_jobs(df, resource):
    completed_df = load_data_completed_jobs(resource)

    if completed_df.empty:
        return df

    completed_job_numbers = completed_df["Number"].dropna().unique()
    df = df[~df["Number"].isin(completed_job_numbers)]

    return df

# Maps each schedule key to the Statii column and value used to identify completed jobs.
# "col" is the Statii scheduling report column to filter on ("resource" or "operation").
# "value" is the string to match (case-insensitive, partial match).
STATII_FILTER_MAP = {
    "saw":           {"col": "resource",  "value": "Saw"},
    "weld":          {"col": "resource",  "value": "Welding"},
    "machine":       {"col": "resource",  "value": "Machining"},
    "rubber lining": {"col": "operation", "value": "Rubber lining"},
    # operation values confirmed from /report/scheduling API — verify tube/fold if names differ
    "flat":          {"col": "operation", "value": "Laser - Flat"},
    "tube":          {"col": "operation", "value": "Laser - Tube"},
    "fold":          {"col": "operation", "value": "Brake Press"},
}

def remove_completed_jobs_statii(df, resource):
    completed_data = statii_completed_jobs()
    if not completed_data or not completed_data.get("rows"):
        return df
    completed_df = pd.DataFrame(completed_data["rows"], columns=completed_data["columns"])
    if "number" not in completed_df.columns:
        return df
    filter_cfg = STATII_FILTER_MAP.get(resource)
    if filter_cfg:
        col, value = filter_cfg["col"], filter_cfg["value"]
        if col in completed_df.columns:
            completed_df = completed_df[completed_df[col].str.contains(value, case=False, na=False)]
    completed_numbers = set(completed_df["number"].dropna().astype(str))
    return df[~df["Number"].astype(str).isin(completed_numbers)]

def format_hours(hours):
    if hours != hours:  # NaN check
        return "NO DATA"
    h = int(hours)
    m = int(round((hours - h) * 60))
    return f"{h}h {m}m"